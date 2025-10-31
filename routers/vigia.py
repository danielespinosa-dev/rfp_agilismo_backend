from fastapi import APIRouter, HTTPException, UploadFile, File, Form
from flask import json
from pydantic import BaseModel, Field
from typing import List, Optional
from datetime import datetime
from motor.motor_asyncio import AsyncIOMotorClient
from bson import ObjectId
from io import BytesIO, StringIO
from models import TipoAsistenteEnum
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from dotenv import load_dotenv
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from fastapi import UploadFile
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

import zipfile
import rarfile
import tempfile
import os
import asyncio
import pandas as pd

from services.openai_assistant  import OpenAIAssistant

# Cargar variables de entorno
load_dotenv()

# --- Configuración MongoDB ---
MONGO_URL = os.getenv("MONGO_URL")
client = AsyncIOMotorClient(MONGO_URL)
db = client["RFPScrum"]

# --- Modelos Pydantic ---
class SolicitudModel(BaseModel):
    SolicitudID: Optional[str] = Field(default_factory=lambda: str(ObjectId()))
    CodigoProyecto: str
    ProveedorNombre: str
    ProveedorNIT: str
    FechaCreacion: datetime
    EstadoGeneral: str
    UsuarioSolicitante: str
    FuenteExcelPath: Optional[str] = None
    Anexos: List[dict] = Field(default_factory=list)
    StorageFolderPath: Optional[str] = None
    PuntajeConsolidado: Optional[float] = None
    NivelGlobal: Optional[str] = None
    FechaFinalizacion: Optional[datetime] = None
    Evaluacion: Optional[List[dict]] = None
    Respuesta: Optional[str] = None
    Cuestionario: Optional[str] = None
    Analisis: Optional[str] = None
    Mensaje: Optional[str] = None
    class Config:
        from_attributes = True  # Pydantic v2

async def descomprimir_anexos_recursivo(anexos: list) -> list:
    archivos_finales = []

    def leer_fileobj(fileobj):
        return fileobj.read()

    async def procesar_archivo(fileobj, filename):
        ext = filename.lower().split('.')[-1]
        if ext == 'zip':
            with tempfile.TemporaryDirectory() as tmpdir:
                path = os.path.join(tmpdir, filename)
                os.makedirs(os.path.dirname(path), exist_ok=True)
                with open(path, "wb") as f:
                    f.write(leer_fileobj(fileobj))
                with zipfile.ZipFile(path, 'r') as zf:
                    for info in zf.infolist():
                        if info.is_dir():
                            continue
                        inner_name = info.filename
                        inner_bytes = zf.read(inner_name)
                        if inner_name.lower().endswith(('.zip', '.rar')):
                            await procesar_archivo(BytesIO(inner_bytes), inner_name)
                        else:
                            archivos_finales.append(
                                UploadFile(file=BytesIO(inner_bytes), filename=os.path.basename(inner_name))
                            )
        elif ext == 'rar':
            with tempfile.TemporaryDirectory() as tmpdir:
                path = os.path.join(tmpdir, filename)
                os.makedirs(os.path.dirname(path), exist_ok=True)
                with open(path, "wb") as f:
                    f.write(leer_fileobj(fileobj))
                with rarfile.RarFile(path) as rf:
                    for info in rf.infolist():
                        if info.is_dir():
                            continue
                        inner_name = info.filename
                        inner_bytes = rf.read(inner_name)
                        if inner_name.lower().endswith(('.zip', '.rar')):
                            await procesar_archivo(BytesIO(inner_bytes), inner_name)
                        else:
                            archivos_finales.append(
                                UploadFile(file=BytesIO(inner_bytes), filename=os.path.basename(inner_name))
                            )
        else:
            archivos_finales.append(
                UploadFile(file=BytesIO(leer_fileobj(fileobj)), filename=filename)
            )

    for anexo in anexos:
        await procesar_archivo(anexo.file, anexo.filename)

    return archivos_finales

def extraer_hojas_excel_plano(excel_file) -> str:
    """
    Extrae y concatena en texto plano la información de las siguientes hojas:
    1. Datos del proveedor
    2. Roles
    3. CV . Equipo de trabajo
    4. Propuesta económica
    5. Experiencia
    6. Req del servicio

    Maneja celdas combinadas y retorna un string legible y coherente.
    """
    hojas_objetivo = [
        "1.Datos del proveedor",
        "2. Roles",
        "2.Roles",
        "3. CV . Equipo de trabajo",
        "3.CV-Equipo de trabajo",
        "4. Propuesta económica",
        "5. Experiencia",
        "5. Experiencia Certificaciones",
        "6. Req del servicio"
    ]
    contents = excel_file.file.read()
    wb = load_workbook(filename=BytesIO(contents), data_only=True)
    resultado = []

    for hoja in hojas_objetivo:
        if hoja not in wb.sheetnames:
            resultado.append(f"--- {hoja} ---\n[Hoja no encontrada]\n")
            continue
        ws = wb[hoja]
        # Crear una matriz con los valores de las celdas, rellenando celdas combinadas
        data = []
        for row in ws.iter_rows():
            data_row = []
            for cell in row:
                if isinstance(cell, MergedCell):
                    # Buscar el valor de la celda superior izquierda del rango combinado
                    valor = None
                    for merged_range in ws.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            valor = ws.cell(merged_range.min_row, merged_range.min_col).value
                            break
                    data_row.append(valor)
                else:
                    data_row.append(cell.value)
            data.append(data_row)
        df = pd.DataFrame(data)
        df.dropna(axis=0, how='all', inplace=True)
        df.dropna(axis=1, how='all', inplace=True)
        df = df.fillna("")

        texto_hoja = f"--- {hoja} ---\n"
        if df.shape[0] > 1 and all(isinstance(x, str) for x in df.iloc[0]):
            encabezados = [str(x).strip() for x in df.iloc[0]]
            texto_hoja += " | ".join(encabezados) + "\n"
            texto_hoja += "-" * (len(texto_hoja) - 1) + "\n"
            for _, row in df.iloc[1:].iterrows():
                fila = [str(x).strip() for x in row]
                texto_hoja += " | ".join(fila) + "\n"
        else:
            for _, row in df.iterrows():
                fila = [str(x).strip() for x in row]
                texto_hoja += " | ".join(fila) + "\n"
        resultado.append(texto_hoja)

    return "\n".join(resultado)

def extraer_hojas_excel_json(excel_file: UploadFile) -> dict:
    """
    Extrae y retorna en formato JSON estructurado la información de las siguientes hojas:
    1. Datos del proveedor
    2. Roles
    3. CV . Equipo de trabajo
    4. Propuesta económica
    5. Experiencia
    6. Req del servicio

    Maneja celdas combinadas y retorna una estructura tipo:
    {
        "1.Datos del proveedor": [ {col1: val1, col2: val2, ...}, ... ],
        ...
    }
    """
    from openpyxl.cell.cell import MergedCell

    hojas_objetivo = [
        "1.Datos del proveedor",
        "2. Roles",
        "3. CV . Equipo de trabajo",
        "4. Propuesta económica",
        "5. Experiencia",
        "6. Req del servicio"
    ]
    contents = excel_file.file.read()
    wb = load_workbook(filename=BytesIO(contents), data_only=True)
    resultado = {}

    for hoja in hojas_objetivo:
        if hoja not in wb.sheetnames:
            resultado[hoja] = []
            continue
        ws = wb[hoja]
        # Rellenar celdas combinadas
        for merged in ws.merged_cells.ranges:
            valor = ws.cell(merged.min_row, merged.min_col).value
            for row in ws.iter_rows(merged.min_row, merged.max_row, merged.min_col, merged.max_col):
                for cell in row:
                    if not isinstance(cell, MergedCell):
                        cell.value = valor

        data = list(ws.values)
        # Eliminar filas y columnas completamente vacías
        df = pd.DataFrame(data)
        df.dropna(axis=0, how='all', inplace=True)
        df.dropna(axis=1, how='all', inplace=True)
        df = df.fillna("")

        if df.shape[0] < 2:
            resultado[hoja] = []
            continue

        encabezados = [str(x).strip() for x in df.iloc[0]]
        filas = []
        for _, row in df.iloc[1:].iterrows():
            fila = {col: str(val).strip() for col, val in zip(encabezados, row.values)}
            filas.append(fila)
        resultado[hoja] = filas

    return resultado

def extraer_hojas_excel_a_pdfs(excel_file: UploadFile, output_dir: str = "pdfs_generados") -> list:
    """
    Recibe un archivo Excel y guarda un PDF por cada hoja en la ruta indicada,
    con formato de tabla, colores y márgenes adecuados.
    Retorna una lista con las rutas de los PDFs generados.
    """
    import os
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet

    contents = excel_file.file.read()
    xls = pd.ExcelFile(BytesIO(contents))
    pdf_paths = []

    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    styles = getSampleStyleSheet()
    page_width, page_height = letter
    left_margin = right_margin = 30
    top_margin = bottom_margin = 30
    usable_width = page_width - left_margin - right_margin

    for sheet_name in xls.sheet_names:
        df = xls.parse(sheet_name)
        df = df.fillna("")
        data = [list(df.columns)] + df.values.tolist()

        # Calcular anchos de columna
        num_cols = len(data[0])
        if num_cols == 0:
            continue
        col_width = usable_width / num_cols
        col_widths = [col_width] * num_cols

        safe_sheet_name = "".join([c if c.isalnum() else "_" for c in sheet_name])
        pdf_path = os.path.join(output_dir, f"{safe_sheet_name}.pdf")
        doc = SimpleDocTemplate(
            pdf_path,
            pagesize=letter,
            leftMargin=left_margin,
            rightMargin=right_margin,
            topMargin=top_margin,
            bottomMargin=bottom_margin
        )
        elements = []

        # Título
        elements.append(Paragraph(f"Hoja: {sheet_name}", styles['Title']))
        elements.append(Spacer(1, 12))

        # Tabla
        table = Table(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#dbe5f1")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor("#17375d")),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f2f2")]),
            ('LEFTPADDING', (0, 0), (-1, -1), 4),
            ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)
        doc.build(elements)
        pdf_paths.append(pdf_path)
    return pdf_paths

def extraer_excel_para_assistant(excel_file: UploadFile) -> str:
    """
    Extrae todas las hojas del Excel y las convierte en texto plano tipo tabla,
    omitiendo filas completamente vacías.
    """
    contents = excel_file.file.read()
    xls = pd.ExcelFile(BytesIO(contents))
    resultado = []

    for sheet_name in xls.sheet_names:
        df = xls.parse(sheet_name)
        df = df.fillna("")
        texto_hoja = f"\n=== Hoja: {sheet_name} ===\n"
        headers = [str(h).strip() for h in df.columns]
        texto_hoja += " | ".join(headers) + "\n"
        texto_hoja += "-" * (len(texto_hoja) - 1) + "\n"
        for _, row in df.iterrows():
            fila = [str(x).strip() for x in row]
            # Omitir filas donde todos los valores están vacíos
            if any(fila):
                texto_hoja += " | ".join(fila) + "\n"
        resultado.append(texto_hoja)

    return "\n".join(resultado)

async def procesar_solicitud_con_assistant(
    solicitud: SolicitudModel,
    anexos_ids: list,
    assistant: OpenAIAssistant,
    tipo_asistente: TipoAsistenteEnum
):
    cuestionario=solicitud.Cuestionario
    
    # Formatear anexos para el mensaje
    if anexos_ids:
        anexos_str = ', '.join([f"{a['filename']} (ID: {a['id']})" for a in anexos_ids])
    else:
        anexos_str = 'Ninguno'
    mensaje = (
        f"Solicitud creada para el proyecto {solicitud.CodigoProyecto}.\n"
        f"Proveedor: {solicitud.ProveedorNombre} (NIT: {solicitud.ProveedorNIT}).\n"
        f"Anexos: {anexos_str}.\n"
        f"Datos del formulario diligenciados por proveedor: {cuestionario if cuestionario else 'No hay datos de formulario.'}\n"  
    )
    # print(f"[Vigia] Mensaje Assistant: {mensaje}")
    solicitud.Mensaje = mensaje
    await db.Solicitud.update_one({"SolicitudID": solicitud.SolicitudID}, {"$set": solicitud.dict()})
    doc = await db.Solicitud.find_one({"SolicitudID": solicitud.SolicitudID})
    solicitud = SolicitudModel(**doc) 
    
    max_retries = 3
    retries = 0
    required_actions = []
    current_message = mensaje
    # Solo IDs para assistant
    current_file_ids = [a["id"] for a in anexos_ids]

    while retries < max_retries:
        required_action = await assistant.run_assistant_flow(
            current_message,
            file_ids=current_file_ids,
            tipo_asistente=tipo_asistente
        )
        if required_action:
            required_actions.append(required_action)
            break
        current_message = (
            f"{mensaje}\n\nPor favor, responde ejecutando la función configurada en el assistant. Intento {retries+2}."
        )
        retries += 1

    solicitud.Evaluacion = required_actions
    solicitud.Respuesta = next(
            (ra["assistant_response"] for ra in required_actions if isinstance(ra, dict) and ra.get("assistant_response")), ""
        )
    solicitud.EstadoGeneral = "done" if required_actions else "failed"
    await db.Solicitud.update_one({"SolicitudID": solicitud.SolicitudID}, {"$set": solicitud.dict()})
    doc = await db.Solicitud.find_one({"SolicitudID": solicitud.SolicitudID})
    solicitud = SolicitudModel(**doc) 
    #await assistant.depureFilesV2(current_file_ids)
    OPENAI_VECTOR_STORAGE_ID = os.getenv("OPENAI_VECTOR_STORAGE_ID")
    await assistant.delete_all_files_from_vector_store(vector_store_id=OPENAI_VECTOR_STORAGE_ID)
    await assistant.depureFiles()
    print(f"[Vigia] Solicitud {solicitud.SolicitudID} actualizada tras evaluación")

# --- Router FastAPI ---
router = APIRouter(prefix="/vigia", tags=["Vigia"])

@router.post("/solicitud", response_model=SolicitudModel)
async def create_solicitud(
    CodigoProyecto: str = Form(...),
    ProveedorNombre: str = Form(...),
    ProveedorNIT: str = Form(...),
    EstadoGeneral: str = Form(...),
    UsuarioSolicitante: str = Form(...),
    excel_file: UploadFile = File(...),
    anexos: List[UploadFile] = File(None)
):
    # Configuración del assistant
    # AZURE_OPENAI_API_KEY = os.getenv("AZURE_OPENAI_API_KEY")
    # AZURE_OPENAI_ASSISTANT_ID = os.getenv("AZURE_OPENAI_ASSISTANT_ID")
    # AZURE_OPENAI_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT")
    # AZURE_OPENAI_API_VERSION = os.getenv("AZURE_OPENAI_API_VERSION")
    # assistant = AzureOpenAIAssistant(api_key=AZURE_OPENAI_API_KEY
    #                                  , endpoint=AZURE_OPENAI_ENDPOINT
    #                                  , assistant_id=AZURE_OPENAI_ASSISTANT_ID
    #                                  , api_version=AZURE_OPENAI_API_VERSION)

    OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
    OPENAI_ASSISTANT_ID = os.getenv("OPENAI_ASSISTANT_ID")
    OPENAI_VECTOR_STORAGE_ID = os.getenv("OPENAI_VECTOR_STORAGE_ID")
    assistant = OpenAIAssistant(api_key=OPENAI_API_KEY
                                     , assistant_id= OPENAI_ASSISTANT_ID)
    await assistant.depureFiles()
    #return
    # Extraer cuestionario del Excel
    # cuestionario_csv = extraer_hojas_excel_json(excel_file)
    # cuestionario_csv = extraer_hojas_excel_plano(excel_file)
    # rutas_pdfs = extraer_hojas_excel_a_pdfs(excel_file, output_dir="C:/ruta/deseada")
    cuestionario_csv = extraer_excel_para_assistant(excel_file)
    anexos_descomprimidos = await descomprimir_anexos_recursivo(anexos)
    # Excluir archivos de imagen
    image_extensions = ('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.webp')
    anexos_descomprimidos = [
        anexo for anexo in anexos_descomprimidos
        if not anexo.filename.lower().endswith(image_extensions)
    ]
    # Subir anexos y obtener sus IDs y nombres
    anexos_ids = []
    # anexo_upload = await assistant.upload_file_from_formdata_v2(excel_file, excel_file.filename)
    # if anexo_upload:
    #    anexos_ids.append({"id": anexo_upload["id"], "filename": excel_file.filename})

    if anexos_descomprimidos:
        for anexo in anexos_descomprimidos:
            anexo_upload = await assistant.upload_file_from_formdata_v2(anexo, anexo.filename)
            if anexo_upload:
                anexos_ids.append({"id": anexo_upload["id"], "filename": anexo.filename})
    file_id_list = [anexo["id"] for anexo in anexos_ids if "id" in anexo and anexo["id"]]
    await assistant.add_files_to_vector_store(vector_store_id=OPENAI_VECTOR_STORAGE_ID, file_ids=file_id_list)
    print("[Vigia] Esperando 2 minutos para que el vector store procese los archivos...")
    await asyncio.sleep(120)  # Delay de 2 minutos
    solicitud = SolicitudModel(
        CodigoProyecto=CodigoProyecto,
        ProveedorNombre=ProveedorNombre,
        ProveedorNIT=ProveedorNIT,
        FechaCreacion=datetime.utcnow(),
        EstadoGeneral="En progreso",
        UsuarioSolicitante=UsuarioSolicitante,
        FuenteExcelPath=excel_file.filename,
        Anexos=anexos_ids,
        #Cuestionario=json.dumps(cuestionario_csv, ensure_ascii=False),
        Cuestionario=cuestionario_csv
    )

    await db.Solicitud.insert_one(solicitud.dict())
    print(f"[Vigia] Solicitud creada con ID: {solicitud.SolicitudID}")

    # Procesar los asistentes de forma asíncrona
    asyncio.create_task(procesar_solicitud_con_assistant(solicitud, anexos_ids, assistant, TipoAsistenteEnum.ambiental))

    return solicitud

@router.get("/solicitud/{solicitud_id}", response_model=SolicitudModel)
async def get_solicitud(solicitud_id: str):
    doc = await db.Solicitud.find_one({"SolicitudID": solicitud_id})
    if not doc:
        raise HTTPException(status_code=404, detail="Solicitud not found")
    return SolicitudModel(**doc)

@router.get("/solicitudes", response_model=List[SolicitudModel])
async def list_solicitudes():
    solicitudes = []
    async for doc in db.Solicitud.find():
        solicitudes.append(SolicitudModel(**doc))
    return solicitudes

@router.put("/solicitud/{solicitud_id}", response_model=SolicitudModel)
async def update_solicitud(solicitud_id: str, solicitud: SolicitudModel):
    result = await db.Solicitud.replace_one({"SolicitudID": solicitud_id}, solicitud.dict())
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Solicitud not found")
    return solicitud

@router.delete("/solicitud/{solicitud_id}")
async def delete_solicitud(solicitud_id: str):
    result = await db.Solicitud.delete_one({"SolicitudID": solicitud_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Solicitud not found")
    return {"detail": "Solicitud deleted"}